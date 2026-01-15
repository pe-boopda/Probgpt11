# backend/app/services/export_service.py
"""
Service for exporting test results to Excel and PDF formats
"""

from sqlalchemy.orm import Session
from typing import List, Dict, Any
from datetime import datetime
from io import BytesIO
import logging

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF export
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from app.models.session import TestSession
from app.models.test import Test
from app.models.user import User
from app.services.session_service import SessionService


logger = logging.getLogger(__name__)


class ExportService:
    def __init__(self, db: Session):
        self.db = db
        self.session_service = SessionService(db)

    # ==================== EXCEL EXPORT ====================

    def export_test_results_to_excel(
        self, test_id: int, include_details: bool = False
    ) -> BytesIO:
        """
        Export test results to Excel file
        
        Args:
            test_id: Test ID
            include_details: Include detailed question breakdown
            
        Returns:
            BytesIO: Excel file in memory
        """
        logger.info(f"Exporting test {test_id} results to Excel")
        
        # Get test and sessions
        test = self.db.query(Test).filter(Test.id == test_id).first()
        if not test:
            raise ValueError("Test not found")
        
        sessions = (
            self.db.query(TestSession)
            .filter(
                TestSession.test_id == test_id,
                TestSession.status == "completed"
            )
            .all()
        )
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Сводка"
        self._create_excel_summary_sheet(ws_summary, test, sessions)
        
        # Sheet 2: Detailed Results
        ws_detailed = wb.create_sheet("Детальные результаты")
        self._create_excel_detailed_sheet(ws_detailed, test, sessions)
        
        # Sheet 3: Statistics
        ws_stats = wb.create_sheet("Статистика")
        self._create_excel_statistics_sheet(ws_stats, test, sessions)
        
        # Sheet 4: Question Breakdown (if requested)
        if include_details:
            ws_questions = wb.create_sheet("Анализ вопросов")
            self._create_excel_questions_sheet(ws_questions, test, sessions)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    def _create_excel_summary_sheet(
        self, ws, test: Test, sessions: List[TestSession]
    ):
        """Create summary sheet in Excel"""
        # Title
        ws['A1'] = 'ОТЧЁТ ПО РЕЗУЛЬТАТАМ ТЕСТИРОВАНИЯ'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Test info
        ws['A3'] = 'Название теста:'
        ws['B3'] = test.title
        ws['A4'] = 'Описание:'
        ws['B4'] = test.description or 'Нет описания'
        ws['A5'] = 'Дата создания:'
        ws['B5'] = test.created_at.strftime('%d.%m.%Y %H:%M')
        ws['A6'] = 'Проходной балл:'
        ws['B6'] = f"{test.passing_score}%"
        
        # Statistics
        total = len(sessions)
        passed = sum(1 for s in sessions if s.passed)
        failed = total - passed
        avg_score = sum(s.percentage for s in sessions) / total if total > 0 else 0
        
        ws['A8'] = 'ОБЩАЯ СТАТИСТИКА'
        ws['A8'].font = Font(size=14, bold=True)
        
        ws['A10'] = 'Всего попыток:'
        ws['B10'] = total
        ws['A11'] = 'Сдали:'
        ws['B11'] = passed
        ws['A12'] = 'Не сдали:'
        ws['B12'] = failed
        ws['A13'] = 'Процент сдачи:'
        ws['B13'] = f"{(passed / total * 100) if total > 0 else 0:.1f}%"
        ws['A14'] = 'Средний балл:'
        ws['B14'] = f"{avg_score:.1f}%"
        
        # Style
        for row in range(10, 15):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=12)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_excel_detailed_sheet(
        self, ws, test: Test, sessions: List[TestSession]
    ):
        """Create detailed results sheet in Excel"""
        # Headers
        headers = [
            'ФИО студента',
            'Email',
            'Дата прохождения',
            'Время (мин)',
            'Балл',
            'Макс. балл',
            'Процент',
            'Статус'
        ]
        
        # Header row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row, session in enumerate(sessions, start=2):
            student = self.db.query(User).filter(User.id == session.student_id).first()
            
            ws.cell(row=row, column=1).value = student.full_name if student else "Unknown"
            ws.cell(row=row, column=2).value = student.email if student else ""
            ws.cell(row=row, column=3).value = session.completed_at.strftime('%d.%m.%Y %H:%M')
            ws.cell(row=row, column=4).value = session.time_taken // 60 if session.time_taken else 0
            ws.cell(row=row, column=5).value = session.score
            ws.cell(row=row, column=6).value = session.max_score
            ws.cell(row=row, column=7).value = f"{session.percentage:.1f}%"
            
            # Status with color
            status_cell = ws.cell(row=row, column=8)
            if session.passed:
                status_cell.value = "✓ Сдал"
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            else:
                status_cell.value = "✗ Не сдал"
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(sessions) + 2):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15

    def _create_excel_statistics_sheet(
        self, ws, test: Test, sessions: List[TestSession]
    ):
        """Create statistics sheet in Excel"""
        if not sessions:
            ws['A1'] = 'Нет данных для анализа'
            return
        
        # Score distribution
        ws['A1'] = 'РАСПРЕДЕЛЕНИЕ БАЛЛОВ'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Create bins
        bins = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 100)]
        distribution = {f"{b[0]}-{b[1]}%": 0 for b in bins}
        
        for session in sessions:
            for low, high in bins:
                if low <= session.percentage <= high:
                    distribution[f"{low}-{high}%"] += 1
                    break
        
        row = 3
        ws['A3'] = 'Диапазон'
        ws['B3'] = 'Количество студентов'
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        for label, count in distribution.items():
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count
        
        # Time statistics
        ws['A8'] = 'СТАТИСТИКА ПО ВРЕМЕНИ'
        ws['A8'].font = Font(size=14, bold=True)
        
        times = [s.time_taken for s in sessions if s.time_taken]
        if times:
            avg_time = sum(times) / len(times)
            min_time = min(times)
            max_time = max(times)
            
            ws['A10'] = 'Среднее время:'
            ws['B10'] = f"{avg_time // 60} мин {avg_time % 60} сек"
            ws['A11'] = 'Минимальное время:'
            ws['B11'] = f"{min_time // 60} мин {min_time % 60} сек"
            ws['A12'] = 'Максимальное время:'
            ws['B12'] = f"{max_time // 60} мин {max_time % 60} сек"
        
        # Adjust widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25

    def _create_excel_questions_sheet(
        self, ws, test: Test, sessions: List[TestSession]
    ):
        """Create question breakdown sheet in Excel"""
        ws['A1'] = 'АНАЛИЗ ВОПРОСОВ'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['№', 'Вопрос', 'Тип', 'Правильных ответов', 'Процент', 'Сложность']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Analyze each question
        for idx, question in enumerate(test.questions, start=1):
            correct_count = 0
            total_count = 0
            
            for session in sessions:
                result = self.session_service.get_session_result(session.id)
                for q_result in result.questions:
                    if q_result.question_id == question.id:
                        total_count += 1
                        if q_result.is_correct:
                            correct_count += 1
            
            percentage = (correct_count / total_count * 100) if total_count > 0 else 0
            
            # Difficulty
            if percentage >= 80:
                difficulty = "Легкий"
                color = "C6EFCE"
            elif percentage >= 50:
                difficulty = "Средний"
                color = "FFEB9C"
            else:
                difficulty = "Сложный"
                color = "FFC7CE"
            
            row = 3 + idx
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = question.question_text[:50] + "..."
            ws.cell(row=row, column=3).value = question.question_type
            ws.cell(row=row, column=4).value = f"{correct_count}/{total_count}"
            ws.cell(row=row, column=5).value = f"{percentage:.1f}%"
            
            diff_cell = ws.cell(row=row, column=6)
            diff_cell.value = difficulty
            diff_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Adjust widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

    # ==================== PDF EXPORT ====================

    def export_test_results_to_pdf(
        self, test_id: int, include_details: bool = False
    ) -> BytesIO:
        """
        Export test results to PDF file
        
        Args:
            test_id: Test ID
            include_details: Include detailed question breakdown
            
        Returns:
            BytesIO: PDF file in memory
        """
        logger.info(f"Exporting test {test_id} results to PDF")
        
        # Get test and sessions
        test = self.db.query(Test).filter(Test.id == test_id).first()
        if not test:
            raise ValueError("Test not found")
        
        sessions = (
            self.db.query(TestSession)
            .filter(
                TestSession.test_id == test_id,
                TestSession.status == "completed"
            )
            .all()
        )
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        # Title
        story.append(Paragraph("ОТЧЁТ ПО РЕЗУЛЬТАТАМ ТЕСТИРОВАНИЯ", title_style))
        story.append(Spacer(1, 0.3 * inch))
        
        # Test info
        story.append(Paragraph("Информация о тесте", heading_style))
        test_info = [
            ['Название:', test.title],
            ['Описание:', test.description or 'Нет описания'],
            ['Дата создания:', test.created_at.strftime('%d.%m.%Y %H:%M')],
            ['Проходной балл:', f"{test.passing_score}%"],
        ]
        
        test_table = Table(test_info, colWidths=[2*inch, 4*inch])
        test_table.setStyle(TableStyle([
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(test_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Statistics
        story.append(Paragraph("Общая статистика", heading_style))
        
        total = len(sessions)
        passed = sum(1 for s in sessions if s.passed)
        failed = total - passed
        avg_score = sum(s.percentage for s in sessions) / total if total > 0 else 0
        
        stats_data = [
            ['Метрика', 'Значение'],
            ['Всего попыток', str(total)],
            ['Сдали', str(passed)],
            ['Не сдали', str(failed)],
            ['Процент сдачи', f"{(passed / total * 100) if total > 0 else 0:.1f}%"],
            ['Средний балл', f"{avg_score:.1f}%"],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(stats_table)
        
        # Page break
        story.append(PageBreak())
        
        # Detailed results
        story.append(Paragraph("Детальные результаты", heading_style))
        
        results_data = [['ФИО', 'Email', 'Дата', 'Балл', 'Статус']]
        
        for session in sessions:
            student = self.db.query(User).filter(User.id == session.student_id).first()
            status = "✓ Сдал" if session.passed else "✗ Не сдал"
            
            results_data.append([
                student.full_name if student else "Unknown",
                student.email if student else "",
                session.completed_at.strftime('%d.%m.%Y'),
                f"{session.percentage:.1f}%",
                status
            ])
        
        results_table = Table(results_data, colWidths=[1.5*inch, 2*inch, 1*inch, 1*inch, 1*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(results_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        return output

    # ==================== INDIVIDUAL STUDENT EXPORT ====================

    def export_student_certificate_pdf(self, session_id: int) -> BytesIO:
        """
        Export individual student certificate/report as PDF
        
        Args:
            session_id: Session ID
            
        Returns:
            BytesIO: PDF certificate
        """
        session = self.session_service.get_session(session_id)
        if session.status != "completed":
            raise ValueError("Session is not completed")
        
        result = self.session_service.get_session_result(session_id)
        test = self.db.query(Test).filter(Test.id == session.test_id).first()
        student = self.db.query(User).filter(User.id == session.student_id).first()
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Certificate title
        title_style = ParagraphStyle(
            'CertTitle',
            parent=styles['Title'],
            fontSize=28,
            textColor=colors.HexColor('#1e40af'),
            alignment=TA_CENTER,
            spaceAfter=30
        )
        
        story.append(Spacer(1, 1*inch))
        story.append(Paragraph("СЕРТИФИКАТ О ПРОХОЖДЕНИИ ТЕСТА", title_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Student info
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        story.append(Paragraph(f"<b>{student.full_name}</b>", info_style))
        story.append(Paragraph(f"успешно прошёл(а) тест", info_style))
        story.append(Paragraph(f"<b>{test.title}</b>", info_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Results box
        results_data = [
            ['Результат:', f"{result.percentage:.1f}%"],
            ['Балл:', f"{result.score} из {result.max_score}"],
            ['Дата:', session.completed_at.strftime('%d.%m.%Y')],
            ['Статус:', '✓ Тест сдан' if session.passed else '✗ Тест не сдан'],
        ]
        
        results_table = Table(results_data, colWidths=[2*inch, 3*inch])
        results_table.setStyle(TableStyle([
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 12),
            ('FONT', (1, 0), (1, -1), 'Helvetica', 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#3b82f6')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        story.append(results_table)
        
        # Build
        doc.build(story)
        output.seek(0)
        
        return output